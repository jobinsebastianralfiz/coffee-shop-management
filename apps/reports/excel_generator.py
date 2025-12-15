"""
Excel report generation using openpyxl.
"""

from io import BytesIO
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

from django.utils import timezone

from apps.core.models import BusinessSettings
from apps.orders.models import Order


class BaseExcelReport:
    """Base class for Excel report generation."""

    def __init__(self, title):
        """Initialize Excel report."""
        self.title = title
        self.wb = Workbook()
        self.business = BusinessSettings.load()
        self.currency_symbol = self.business.currency_symbol or "Rs."

        # Define styles
        self._setup_styles()

    def _setup_styles(self):
        """Set up reusable styles."""
        # Header style
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Title style
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(size=12, color="666666")

        # Currency style
        self.currency_font = Font(size=10)

        # Border style
        thin_border = Side(style='thin', color='E5E7EB')
        self.cell_border = Border(
            left=thin_border, right=thin_border,
            top=thin_border, bottom=thin_border
        )

        # Alternate row fill
        self.alt_row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    def create_worksheet(self, name):
        """Create a new worksheet."""
        if self.wb.active and self.wb.active.title == "Sheet":
            ws = self.wb.active
            ws.title = name
        else:
            ws = self.wb.create_sheet(name)
        return ws

    def add_title(self, ws, title, row=1):
        """Add a title to the worksheet."""
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = self.title_font
        return row + 1

    def add_subtitle(self, ws, subtitle, row=2):
        """Add a subtitle to the worksheet."""
        ws.cell(row=row, column=1, value=subtitle)
        ws.cell(row=row, column=1).font = self.subtitle_font
        return row + 1

    def add_header_row(self, ws, headers, row):
        """Add a styled header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border
        return row + 1

    def add_data_row(self, ws, data, row, is_alternate=False):
        """Add a data row."""
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = self.cell_border
            if is_alternate:
                cell.fill = self.alt_row_fill
        return row + 1

    def format_currency(self, value):
        """Format a value as currency."""
        if value is None:
            value = 0
        return f"{self.currency_symbol}{value:,.2f}"

    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths."""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    def generate(self):
        """Generate the Excel file."""
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer


class SalesExcelReport(BaseExcelReport):
    """Sales report Excel generator."""

    def __init__(self, start_date, end_date):
        """Initialize sales report."""
        super().__init__("Sales Report")
        self.start_date = start_date
        self.end_date = end_date

    def generate_report(self, sales_data, items_data, orders_list, payment_data):
        """Generate complete sales report."""
        # Summary Sheet
        self._create_summary_sheet(sales_data, payment_data)

        # Items Sheet
        self._create_items_sheet(items_data)

        # Orders Sheet
        self._create_orders_sheet(orders_list)

        return self.generate()

    def _create_summary_sheet(self, sales_data, payment_data):
        """Create summary worksheet."""
        ws = self.create_worksheet("Summary")

        row = self.add_title(ws, "Sales Summary")
        row = self.add_subtitle(ws, f"{self.start_date.strftime('%d %b %Y')} - {self.end_date.strftime('%d %b %Y')}", row)
        row += 1

        # Summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Orders", sales_data.get('total_orders', 0)],
            ["Gross Revenue", self.format_currency(sales_data.get('total_revenue', 0))],
            ["Discounts", self.format_currency(sales_data.get('total_discount', 0))],
            ["Net Revenue", self.format_currency(sales_data.get('net_revenue', 0))],
            ["Total Tax", self.format_currency(sales_data.get('total_tax', 0))],
            ["Average Order Value", self.format_currency(sales_data.get('avg_order_value', 0))],
        ]

        row = self.add_header_row(ws, summary_data[0], row)
        for i, data in enumerate(summary_data[1:]):
            row = self.add_data_row(ws, data, row, i % 2 == 1)

        # Payment breakdown
        if payment_data:
            row += 2
            ws.cell(row=row, column=1, value="Payment Methods")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1

            row = self.add_header_row(ws, ["Method", "Transactions", "Amount"], row)
            for i, payment in enumerate(payment_data):
                data = [
                    payment['method'].upper(),
                    payment['count'],
                    self.format_currency(payment['total'])
                ]
                row = self.add_data_row(ws, data, row, i % 2 == 1)

        self.auto_adjust_columns(ws)

    def _create_items_sheet(self, items_data):
        """Create items worksheet."""
        ws = self.create_worksheet("Items")

        row = self.add_title(ws, "Item Sales")
        row = self.add_subtitle(ws, f"{self.start_date.strftime('%d %b %Y')} - {self.end_date.strftime('%d %b %Y')}", row)
        row += 1

        headers = ["Item Name", "Category", "Quantity Sold", "Revenue"]
        row = self.add_header_row(ws, headers, row)

        for i, item in enumerate(items_data):
            data = [
                item.get('menu_item__name', 'N/A'),
                item.get('menu_item__category__name', 'N/A'),
                item.get('total_quantity', 0),
                self.format_currency(item.get('total_revenue', 0))
            ]
            row = self.add_data_row(ws, data, row, i % 2 == 1)

        self.auto_adjust_columns(ws)

    def _create_orders_sheet(self, orders_list):
        """Create orders worksheet."""
        ws = self.create_worksheet("Orders")

        row = self.add_title(ws, "Order Details")
        row = self.add_subtitle(ws, f"{self.start_date.strftime('%d %b %Y')} - {self.end_date.strftime('%d %b %Y')}", row)
        row += 1

        headers = ["Order #", "Date", "Time", "Table", "Type", "Items", "Subtotal", "Discount", "Tax", "Total", "Status"]
        row = self.add_header_row(ws, headers, row)

        for i, order in enumerate(orders_list):
            data = [
                order.order_number,
                order.created_at.strftime('%d/%m/%Y'),
                order.created_at.strftime('%H:%M'),
                order.table.number if order.table else "N/A",
                order.order_type.replace('_', ' ').title(),
                order.item_count,
                self.format_currency(order.subtotal),
                self.format_currency(order.discount_amount),
                self.format_currency(order.cgst_amount + order.sgst_amount),
                self.format_currency(order.total_amount),
                order.status.title()
            ]
            row = self.add_data_row(ws, data, row, i % 2 == 1)

        self.auto_adjust_columns(ws)


class FinancialExcelReport(BaseExcelReport):
    """Financial report Excel generator."""

    def __init__(self, start_date, end_date):
        """Initialize financial report."""
        super().__init__("Financial Report")
        self.start_date = start_date
        self.end_date = end_date

    def generate_report(self, revenue_data, expense_data, pnl_data):
        """Generate financial report."""
        # P&L Summary
        self._create_pnl_sheet(pnl_data)

        # Revenue Details
        self._create_revenue_sheet(revenue_data)

        # Expenses
        self._create_expense_sheet(expense_data)

        return self.generate()

    def _create_pnl_sheet(self, pnl_data):
        """Create P&L summary sheet."""
        ws = self.create_worksheet("Profit & Loss")

        row = self.add_title(ws, "Profit & Loss Statement")
        row = self.add_subtitle(ws, f"{self.start_date.strftime('%d %b %Y')} - {self.end_date.strftime('%d %b %Y')}", row)
        row += 1

        revenue = pnl_data.get('revenue', {})
        expenses = pnl_data.get('expenses', {})

        # Revenue section
        data = [
            ["Revenue", ""],
            ["Gross Sales", self.format_currency(revenue.get('subtotal', 0))],
            ["Less: Discounts", self.format_currency(revenue.get('discounts', 0))],
            ["Net Sales", self.format_currency(revenue.get('net_sales', 0))],
            ["", ""],
            ["Expenses", ""],
            ["Total Expenses", self.format_currency(expenses.get('total', 0))],
            ["", ""],
            ["Gross Profit", self.format_currency(pnl_data.get('gross_profit', 0))],
        ]

        row = self.add_header_row(ws, ["Description", "Amount"], row)
        for i, d in enumerate(data):
            row = self.add_data_row(ws, d, row, i % 2 == 1)

        self.auto_adjust_columns(ws)

    def _create_revenue_sheet(self, revenue_data):
        """Create revenue details sheet."""
        ws = self.create_worksheet("Revenue")

        row = self.add_title(ws, "Revenue Breakdown")
        row += 1

        data = [
            ["Component", "Amount"],
            ["Gross Revenue", self.format_currency(revenue_data.get('gross_revenue', 0))],
            ["Subtotal", self.format_currency(revenue_data.get('subtotal', 0))],
            ["Discounts", self.format_currency(revenue_data.get('discounts', 0))],
            ["CGST", self.format_currency(revenue_data.get('cgst', 0))],
            ["SGST", self.format_currency(revenue_data.get('sgst', 0))],
            ["Service Charge", self.format_currency(revenue_data.get('service_charge', 0))],
            ["Net Sales", self.format_currency(revenue_data.get('net_sales', 0))],
        ]

        row = self.add_header_row(ws, data[0], row)
        for i, d in enumerate(data[1:]):
            row = self.add_data_row(ws, d, row, i % 2 == 1)

        self.auto_adjust_columns(ws)

    def _create_expense_sheet(self, expense_data):
        """Create expense breakdown sheet."""
        ws = self.create_worksheet("Expenses")

        row = self.add_title(ws, "Expense Breakdown")
        row += 1

        ws.cell(row=row, column=1, value=f"Total Expenses: {self.format_currency(expense_data.get('total', 0))}")
        row += 2

        if expense_data.get('by_category'):
            row = self.add_header_row(ws, ["Category", "Count", "Amount"], row)
            for i, cat in enumerate(expense_data['by_category']):
                data = [
                    cat.get('category__name', 'Uncategorized'),
                    cat.get('count', 0),
                    self.format_currency(cat.get('total', 0))
                ]
                row = self.add_data_row(ws, data, row, i % 2 == 1)

        self.auto_adjust_columns(ws)


class ItemSalesExcelReport(BaseExcelReport):
    """Item sales detail Excel report."""

    def __init__(self, start_date, end_date):
        """Initialize item sales report."""
        super().__init__("Item Sales Report")
        self.start_date = start_date
        self.end_date = end_date

    def generate_report(self, items_data, category_data):
        """Generate item sales report."""
        # Items sheet
        ws = self.create_worksheet("Items")
        row = self.add_title(ws, "Item-wise Sales")
        row = self.add_subtitle(ws, f"{self.start_date.strftime('%d %b %Y')} - {self.end_date.strftime('%d %b %Y')}", row)
        row += 1

        headers = ["Rank", "Item", "Category", "Qty Sold", "Revenue", "% of Total"]
        row = self.add_header_row(ws, headers, row)

        total_revenue = sum(item.get('total_revenue', 0) for item in items_data) or 1

        for i, item in enumerate(items_data):
            revenue = item.get('total_revenue', 0)
            percentage = (revenue / total_revenue * 100) if total_revenue else 0
            data = [
                i + 1,
                item.get('menu_item__name', 'N/A'),
                item.get('menu_item__category__name', 'N/A'),
                item.get('total_quantity', 0),
                self.format_currency(revenue),
                f"{percentage:.1f}%"
            ]
            row = self.add_data_row(ws, data, row, i % 2 == 1)

        self.auto_adjust_columns(ws)

        # Categories sheet
        ws2 = self.create_worksheet("Categories")
        row = self.add_title(ws2, "Category-wise Sales")
        row += 1

        headers = ["Category", "Items Sold", "Total Qty", "Revenue"]
        row = self.add_header_row(ws2, headers, row)

        for i, cat in enumerate(category_data):
            data = [
                cat.get('menu_item__category__name', 'N/A'),
                cat.get('item_count', 0),
                cat.get('total_quantity', 0),
                self.format_currency(cat.get('total_revenue', 0))
            ]
            row = self.add_data_row(ws2, data, row, i % 2 == 1)

        self.auto_adjust_columns(ws2)

        return self.generate()
